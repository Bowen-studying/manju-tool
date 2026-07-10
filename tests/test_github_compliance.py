import re
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from manju.pipeline.storyboard_schema import normalize_storyboard
from manju.pipeline.video import _generate_video_prompts
from manju.pipeline.voice import _generate_voice_scripts
from manju.utils.formats import HAS_EXCEL, write_xlsx


ROOT = Path(__file__).resolve().parents[1]
PUBLIC_DOCS = [ROOT / "README.md", *(ROOT / "docs").glob("*.md")]
FORBIDDEN_PROVIDER_NAMES = (
    "Deep" + "Seek",
    "G" + "LM",
    "Ag" + "nes",
    "即" + "梦",
    "Comfy" + "UI",
    "Seed" + "ance",
    "Mini" + "max",
    "剪" + "映",
    "Eleven" + "Labs",
)
PUBLIC_DIRECTIVES = ("不" + "要", "禁" + "止")
PRIVACY_NOTICE = (
    "小说、提示词和参考素材会发送给配置的第三方服务。"
    "处理未公开内容前，请先确认服务商的数据保留与隐私政策"
)


def _upload_text_files():
    paths = [ROOT / "README.md", ROOT / "pyproject.toml", ROOT / "requirements.txt"]
    for directory in (ROOT / "manju", ROOT / "tests", ROOT / "docs"):
        paths.extend(path for path in directory.rglob("*") if path.suffix in {".py", ".md"})
    return paths


class PublicContentComplianceTests(unittest.TestCase):
    def test_upload_sources_are_provider_neutral(self):
        matches = []
        for path in _upload_text_files():
            text = path.read_text(encoding="utf-8")
            for name in FORBIDDEN_PROVIDER_NAMES:
                if name.casefold() in text.casefold():
                    matches.append(f"{path.relative_to(ROOT)}: {name}")
            if "8" + "K" in text.upper():
                matches.append(f"{path.relative_to(ROOT)}: disallowed resolution label")
        self.assertEqual(matches, [])

    def test_public_docs_avoid_internal_directives(self):
        matches = []
        for path in PUBLIC_DOCS:
            text = path.read_text(encoding="utf-8")
            for phrase in PUBLIC_DIRECTIVES:
                if phrase in text:
                    matches.append(f"{path.relative_to(ROOT)}: {phrase}")
        self.assertEqual(matches, [])

    def test_readme_uses_generic_config_and_privacy_notice(self):
        readme = (ROOT / "README.md").read_text(encoding="utf-8")
        for variable in ("LLM_API_KEY", "LLM_API_BASE", "LLM_MODEL"):
            self.assertIn(variable, readme)
        self.assertIn(PRIVACY_NOTICE, readme)

    def test_sensitive_local_config_is_ignored(self):
        patterns = (ROOT / ".gitignore").read_text(encoding="utf-8").splitlines()
        self.assertIn(".env", patterns)
        self.assertIn(".manju.env", patterns)

    def test_repository_text_has_no_embedded_credentials(self):
        secret_patterns = [
            re.compile(r"sk-[A-Za-z0-9_-]{12,}"),
            re.compile(r"-----BEGIN (?:RSA |EC |OPENSSH )?PRIVATE KEY-----"),
        ]
        matches = []
        for path in _upload_text_files():
            text = path.read_text(encoding="utf-8")
            for pattern in secret_patterns:
                if pattern.search(text):
                    matches.append(str(path.relative_to(ROOT)))
        self.assertEqual(matches, [])


class OutputRuleComplianceTests(unittest.TestCase):
    def _storyboard(self):
        return normalize_storyboard({
            "title": "测试",
            "style_anchor": "cinematic anime，电影感",
            "scenes": [{
                "scene_id": 1,
                "scene_heading": "EXT. 天台 - 黄昏",
                "shots": [{
                    "shot_id": "1.1",
                    "shot_type": "近景",
                    "composition": "三分法",
                    "camera_movement": "缓慢推进",
                    "visual_description": "角色转身看向门口",
                    "dialogue_narration": "甲：你好",
                    "image_prompt_en": "a person turns toward the door",
                }],
            }],
        })

    @unittest.skipUnless(HAS_EXCEL, "openpyxl not installed")
    def test_storyboard_xlsx_has_required_columns_and_style(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as output_dir:
            path = Path(output_dir) / "storyboard.xlsx"
            write_xlsx(self._storyboard(), str(path), "分镜表")
            sheet = load_workbook(path).active

        self.assertEqual(sheet.max_column, 18)
        header = sheet.cell(1, 1)
        body = sheet.cell(2, 1)
        self.assertEqual(header.font.name, "宋体")
        self.assertFalse(header.font.bold)
        self.assertTrue(header.fill.fgColor.rgb.endswith("D6E4F0"))
        self.assertEqual(body.font.name, "宋体")
        self.assertFalse(body.font.bold)
        self.assertTrue(body.fill.fgColor.rgb.endswith("FFFFFF"))
        self.assertTrue(body.border.left.color.rgb.endswith("B7B7B7"))

    def test_english_video_prompt_contains_no_chinese(self):
        prompt = _generate_video_prompts(self._storyboard())[0]["video_prompt_en"]
        self.assertIsNone(re.search(r"[\u4e00-\u9fff]", prompt))

    def test_voice_rows_cover_silent_shots(self):
        storyboard = self._storyboard()
        storyboard["scenes"][0]["shots"].append({
            "shot_id": "1.2",
            "visual": {"description": "空镜"},
            "audio": {"speaker": "", "dialogue": "", "narration": ""},
            "prompts": {},
            "assets": {},
            "status": {},
        })
        with patch("manju.pipeline.voice._batch_infer_emotions", return_value={1: "平静"}):
            rows = _generate_voice_scripts(storyboard)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1]["text"], "（无对白）")


if __name__ == "__main__":
    unittest.main()
